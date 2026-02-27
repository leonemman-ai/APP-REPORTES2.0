from fastapi import FastAPI, APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
from pathlib import Path
import os
import logging
import uuid
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import pandas as pd
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Directories
UPLOAD_FOLDER = ROOT_DIR / "uploads"
TEMPLATES_FOLDER = ROOT_DIR / "templates"
GENERATED_FOLDER = ROOT_DIR / "generated"

UPLOAD_FOLDER.mkdir(exist_ok=True)
TEMPLATES_FOLDER.mkdir(exist_ok=True)
GENERATED_FOLDER.mkdir(exist_ok=True)

TEMPLATE_FILE = TEMPLATES_FOLDER / "plantilla.xlsx"

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ===================== MODELS =====================

class Afiliacion(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    codigo: str
    municipio: str
    nombre_comercial: str
    direccion: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TroubleTicket(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    folio: str
    servicio: str
    tecnologia: str
    descripcion: str
    afiliacion: str
    fecha: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DocumentoGenerado(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    folio: str
    filename: str
    filepath: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    data: Dict[str, Any] = Field(default_factory=dict)

# ===================== HELPER FUNCTIONS =====================

async def cargar_afiliaciones_desde_excel(file_path: Path):
    """Carga afiliaciones desde archivo Excel a MongoDB"""
    try:
        excel_sheets = pd.read_excel(file_path, sheet_name=None)
        afiliaciones_list = []
        
        for sheet_name, df in excel_sheets.items():
            df.columns = df.columns.str.strip().str.upper()
            
            if "AFILIACIONES" not in df.columns:
                continue
            
            for _, row in df.iterrows():
                afiliacion_codigo = str(row.get("AFILIACIONES", "")).strip()
                
                if afiliacion_codigo == "" or afiliacion_codigo.lower() == "nan":
                    continue
                
                afiliacion = {
                    "id": str(uuid.uuid4()),
                    "codigo": afiliacion_codigo,
                    "municipio": str(row.get("C5 TOLUCA", "")).strip(),
                    "nombre_comercial": str(row.get("NOMBRE COMERCIAL", "")).strip(),
                    "direccion": str(row.get("DIRECCION", "")).strip(),
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                afiliaciones_list.append(afiliacion)
        
        # Limpiar colección y insertar nuevos datos
        await db.afiliaciones.delete_many({})
        if afiliaciones_list:
            await db.afiliaciones.insert_many(afiliaciones_list)
        
        logger.info(f"Cargadas {len(afiliaciones_list)} afiliaciones")
        return len(afiliaciones_list)
    
    except Exception as e:
        logger.error(f"Error cargando afiliaciones: {e}")
        raise HTTPException(status_code=500, detail=str(e))

async def cargar_tt_desde_excel(file_path: Path):
    """Carga trouble tickets desde archivo Excel a MongoDB"""
    try:
        # Leer TODO el archivo Excel sin límite de filas
        logger.info(f"Iniciando carga de TT desde: {file_path}")
        df = pd.read_excel(file_path, sheet_name=0)  # Leer primera hoja
        
        logger.info(f"Total de filas en el archivo: {len(df)}")
        
        tt_list = []
        filas_invalidas = 0
        
        for idx, row in df.iterrows():
            try:
                # Obtener folio y verificar que no esté vacío
                folio = str(row.iloc[0]).strip()
                
                # Ignorar filas sin folio o con valores vacíos/nan
                if not folio or folio.lower() in ['nan', 'none', '']:
                    filas_invalidas += 1
                    continue
                
                tt = {
                    "id": str(uuid.uuid4()),
                    "folio": folio,
                    "servicio": str(row.iloc[2]).strip() if not pd.isna(row.iloc[2]) else "",
                    "tecnologia": str(row.iloc[3]).strip() if not pd.isna(row.iloc[3]) else "",
                    "descripcion": str(row.iloc[10]).strip() if not pd.isna(row.iloc[10]) else "",
                    "afiliacion": str(row.iloc[9]).strip() if not pd.isna(row.iloc[9]) else "",
                    "fecha": str(row.iloc[12]).strip() if not pd.isna(row.iloc[12]) else "",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                tt_list.append(tt)
                
            except Exception as e:
                logger.warning(f"Error procesando fila {idx}: {e}")
                filas_invalidas += 1
                continue
        
        logger.info(f"Procesadas {len(tt_list)} trouble tickets válidos, {filas_invalidas} filas inválidas")
        
        # Limpiar colección y insertar nuevos datos
        await db.trouble_tickets.delete_many({})
        if tt_list:
            # Insertar en lotes de 1000 para evitar problemas con documentos grandes
            batch_size = 1000
            for i in range(0, len(tt_list), batch_size):
                batch = tt_list[i:i + batch_size]
                await db.trouble_tickets.insert_many(batch)
                logger.info(f"Insertados {len(batch)} TT (lote {i//batch_size + 1})")
        
        logger.info(f"✅ Cargados TODOS los {len(tt_list)} trouble tickets exitosamente")
        return len(tt_list)
    
    except Exception as e:
        logger.error(f"❌ Error cargando TT: {e}")
        raise HTTPException(status_code=500, detail=str(e))

def procesar_imagen(file, target_path: Path):
    """Procesa imagen: redimensiona, recorta y guarda como JPEG"""
    try:
        # Guardar archivo temporal
        temp_path = target_path.parent / f"temp_{target_path.name}"
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Abrir y procesar imagen
        img = Image.open(temp_path)
        
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        
        FRAME_WIDTH = 420
        FRAME_HEIGHT = 420
        
        img_width, img_height = img.size
        ratio = max(FRAME_WIDTH / img_width, FRAME_HEIGHT / img_height)
        
        new_width = int(img_width * ratio)
        new_height = int(img_height * ratio)
        
        img = img.resize((new_width, new_height), Image.LANCZOS)
        
        left = (new_width - FRAME_WIDTH) / 2
        top = (new_height - FRAME_HEIGHT) / 2
        right = left + FRAME_WIDTH
        bottom = top + FRAME_HEIGHT
        
        img = img.crop((left, top, right, bottom))
        
        # Guardar como JPEG
        final_path = target_path.parent / f"{target_path.stem}.jpg"
        img.save(final_path, "JPEG", quality=95)
        
        # Limpiar temporal
        temp_path.unlink()
        
        return final_path
    
    except Exception as e:
        logger.error(f"Error procesando imagen: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando imagen: {str(e)}")

# ===================== API ENDPOINTS =====================

@api_router.get("/")
async def root():
    return {"message": "API de Órdenes de Incidente y Requerimiento"}

@api_router.post("/afiliaciones/upload")
async def upload_afiliaciones(file: UploadFile = File(...)):
    """Subir archivo Excel de afiliaciones"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsx")
    
    file_path = TEMPLATES_FOLDER / "afiliaciones_temp.xlsx"
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    count = await cargar_afiliaciones_desde_excel(file_path)
    
    return {"message": f"Cargadas {count} afiliaciones exitosamente"}

@api_router.get("/afiliaciones")
async def get_afiliaciones():
    """Obtener todas las afiliaciones"""
    afiliaciones = await db.afiliaciones.find({}, {"_id": 0}).to_list(10000)
    return afiliaciones

@api_router.get("/afiliaciones/search")
async def search_afiliacion(q: str):
    """Buscar afiliación por código"""
    afiliacion = await db.afiliaciones.find_one({"codigo": q}, {"_id": 0})
    if not afiliacion:
        return None
    return afiliacion

@api_router.post("/tt/upload")
async def upload_tt(file: UploadFile = File(...)):
    """Subir archivo Excel de Trouble Tickets diario"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsx")
    
    file_path = UPLOAD_FOLDER / "tt_diario_temp.xlsx"
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    count = await cargar_tt_desde_excel(file_path)
    
    return {"message": f"Cargados {count} trouble tickets exitosamente"}

@api_router.get("/tt")
async def get_all_tt():
    """Obtener todos los trouble tickets"""
    tts = await db.trouble_tickets.find({}, {"_id": 0}).to_list(10000)
    return tts

@api_router.get("/tt/{folio}")
async def get_tt_by_folio(folio: str):
    """Obtener TT por folio"""
    tt = await db.trouble_tickets.find_one({"folio": folio}, {"_id": 0})
    if not tt:
        return None
    return tt

@api_router.post("/documentos/generar")
async def generar_documento(
    folio: str = Form(...),
    municipio: str = Form(""),
    direccion: str = Form(""),
    afiliacion: str = Form(""),
    nombre_comercial: str = Form(""),
    solicitante: str = Form(""),
    fecha_creacion: str = Form(""),
    atendido_por: str = Form(""),
    tecnologia: str = Form(""),
    servicio: str = Form(""),
    fecha_llegada: str = Form(""),
    cliente_info: str = Form(""),
    descripcion_falla: str = Form(""),
    afiliacion2: str = Form(""),
    falla_declarada: str = Form(""),
    diagnostico_real: str = Form(""),
    estatus_folio: str = Form(""),
    fecha_atencion: str = Form(""),
    diagnostico: str = Form(""),
    acciones: str = Form(""),
    estatus_reparacion: str = Form(""),
    voltaje: str = Form(""),
    razon1: str = Form(""),
    razon2: str = Form(""),
    cliente2: str = Form(""),
    descripcion2: str = Form(""),
    folio2: str = Form(""),
    diagnostico_detallado: str = Form(""),
    solucion: str = Form(""),
    componentes_instalados: str = Form(""),
    componentes_retirados: str = Form(""),
    tecnico: str = Form(""),
    supervisor: str = Form(""),
    antes: Optional[UploadFile] = File(None),
    proceso1: Optional[UploadFile] = File(None),
    proceso2: Optional[UploadFile] = File(None),
    final: Optional[UploadFile] = File(None)
):
    """Generar documento Excel con todos los datos"""
    try:
        # Cargar plantilla
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active
        
        # Obtener datos del TT si existe
        tt = await db.trouble_tickets.find_one({"folio": folio}, {"_id": 0})
        
        if tt:
            servicio = tt.get("servicio", servicio)
            tecnologia = tt.get("tecnologia", tecnologia)
            descripcion_tt = tt.get("descripcion", "")
            afiliacion_tt = tt.get("afiliacion", "")
            fecha_tt = tt.get("fecha", "")
        else:
            descripcion_tt = ""
            afiliacion_tt = ""
            fecha_tt = ""
        
        # Llenar datos en la plantilla
        ws["J19"] = folio
        ws["C21"] = municipio
        ws["F21"] = direccion
        ws["C23"] = afiliacion_tt if afiliacion_tt else afiliacion
        ws["F23"] = nombre_comercial
        ws["C25"] = solicitante
        ws["F25"] = fecha_tt if fecha_tt else fecha_creacion
        
        ws["C34"] = atendido_por
        ws["F34"] = tecnologia
        ws["I34"] = servicio
        ws["C37"] = fecha_llegada
        ws["H37"] = cliente_info
        ws["H38"] = descripcion_tt if descripcion_tt else descripcion_falla
        ws["H39"] = afiliacion_tt if afiliacion_tt else afiliacion2
        ws["C41"] = descripcion_tt if descripcion_tt else falla_declarada
        ws["F41"] = diagnostico_real
        ws["C47"] = estatus_folio
        ws["F45"] = fecha_atencion
        
        ws["F55"] = diagnostico
        ws["F56"] = acciones
        ws["F57"] = estatus_reparacion
        ws["F58"] = voltaje
        ws["F59"] = razon1
        ws["F60"] = razon2
        ws["F61"] = cliente2
        ws["F62"] = descripcion2
        ws["F63"] = folio2
        ws["F64"] = diagnostico_detallado
        ws["F74"] = solucion
        
        ws["C88"] = componentes_instalados
        ws["C91"] = componentes_retirados
        
        ws["C149"] = tecnico
        ws["H149"] = supervisor
        
        # Procesar y añadir imágenes
        imagenes = [
            (antes, "C100"),
            (proceso1, "H100"),
            (proceso2, "C126"),
            (final, "H126")
        ]
        
        for img_file, cell in imagenes:
            if img_file and img_file.filename:
                img_id = str(uuid.uuid4())
                img_path = UPLOAD_FOLDER / f"{img_id}.jpg"
                
                # Procesar imagen
                processed_path = procesar_imagen(img_file, img_path)
                
                # Añadir a Excel
                xl_img = XLImage(str(processed_path))
                xl_img.width = 420
                xl_img.height = 420
                ws.add_image(xl_img, cell)
        
        # Configurar página
        ws.page_setup.fitToWidth = 1
        ws.sheet_view.zoomScale = 100
        
        # Guardar documento
        folio_safe = re.sub(r'[\\/*?:"<>|]', "-", folio if folio else "SIN_FOLIO")
        output_filename = f"{folio_safe}.xlsx"
        output_path = GENERATED_FOLDER / output_filename
        
        wb.save(output_path)
        
        # Guardar registro en DB
        documento = {
            "id": str(uuid.uuid4()),
            "folio": folio,
            "filename": output_filename,
            "filepath": str(output_path),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "data": {
                "municipio": municipio,
                "afiliacion": afiliacion,
                "tecnico": tecnico,
                "supervisor": supervisor
            }
        }
        await db.documentos.insert_one(documento)
        
        return FileResponse(
            path=str(output_path),
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except Exception as e:
        logger.error(f"Error generando documento: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/documentos")
async def get_documentos():
    """Obtener lista de documentos generados"""
    documentos = await db.documentos.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return documentos

@api_router.get("/documentos/{doc_id}/download")
async def download_documento(doc_id: str):
    """Descargar documento generado"""
    documento = await db.documentos.find_one({"id": doc_id}, {"_id": 0})
    
    if not documento:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    filepath = Path(documento["filepath"])
    
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    return FileResponse(
        path=str(filepath),
        filename=documento["filename"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    """Cargar afiliaciones desde archivo base si existe"""
    afiliaciones_file = TEMPLATES_FOLDER / "AFILIACIONES_CASEOLI.xlsx"
    if afiliaciones_file.exists():
        count = await db.afiliaciones.count_documents({})
        if count == 0:
            logger.info("Cargando afiliaciones desde archivo base...")
            await cargar_afiliaciones_desde_excel(afiliaciones_file)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
